import io
from datetime import date
from flask import Blueprint, request, render_template, redirect, flash, url_for, jsonify, session, send_file
from db import (get_db, dict_cur, get_batches_full, get_users, get_registrations, set_registration_status,
                build_batch_label, TYPE_OFFSETS, create_account_token, get_solar_requests,
                audit_log, get_audit_logs, get_feedback, get_feedback_summary, get_analytics_data)
from email_utils import send_acceptance_email

admin_bp = Blueprint("admin", __name__)


@admin_bp.route("/admin")
def admin():
    from flask import session
    if session.get("user_role") != "admin":
        return redirect("/auth")
    users = get_users()
    uc = {
        "admin":          sum(1 for u in users if u["role"] == "admin"),
        "employee":       sum(1 for u in users if u["role"] == "employee"),
        "trainee":        sum(1 for u in users if u["role"] == "trainee"),
        "archived":       sum(1 for u in users if u["archived_at"] is not None),
        "active":         sum(1 for u in users if u["archived_at"] is None),
        "active_trainee": sum(1 for u in users if u["role"] == "trainee" and u["archived_at"] is None),
    }
    from datetime import timedelta
    all_batches = get_batches_full()
    today = date.today()
    end_offsets = {"training": 9, "renewal": 1}
    upcoming, past = [], []
    for b in all_batches:
        b["end_date"] = b["start_date"] + timedelta(days=end_offsets.get(b["type"], 0))
        (past if b["end_date"] < today else upcoming).append(b)

    # Get LMS modules (active only)
    conn = get_db()
    cur = dict_cur(conn)
    cur.execute("""
        SELECT m.*, u.fname, u.lname,
               COUNT(DISTINCT q.id) as quiz_count,
               COUNT(DISTINCT e.id) as exam_count
        FROM lms_modules m
        LEFT JOIN users u ON m.instructor_id = u.id
        LEFT JOIN lms_quizzes q ON m.id = q.module_id
        LEFT JOIN lms_practical_exams e ON m.id = e.module_id
        WHERE m.archived_at IS NULL
        GROUP BY m.id, u.id, u.fname, u.lname
        ORDER BY m.created_at DESC
    """)
    modules = cur.fetchall()

    # Get archived modules
    cur.execute("""
        SELECT m.*, u.fname, u.lname,
               COUNT(DISTINCT q.id) as quiz_count,
               COUNT(DISTINCT e.id) as exam_count
        FROM lms_modules m
        LEFT JOIN users u ON m.instructor_id = u.id
        LEFT JOIN lms_quizzes q ON m.id = q.module_id
        LEFT JOIN lms_practical_exams e ON m.id = e.module_id
        WHERE m.archived_at IS NOT NULL
        GROUP BY m.id, u.id, u.fname, u.lname
        ORDER BY m.archived_at DESC
    """)
    archived_modules = cur.fetchall()
    cur.close()
    conn.close()

    return render_template("admin/admin.html",
                           registrations=get_registrations(),
                           schedules=upcoming,
                           past_schedules=past,
                           users=users,
                           uc=uc,
                           solar_requests=get_solar_requests(),
                           modules=modules,
                           archived_modules=archived_modules)


@admin_bp.route("/admin/user/archive/<int:user_id>", methods=["POST"])
def admin_archive_user(user_id):
    conn = get_db()
    cur  = dict_cur(conn)
    cur.execute("SELECT fname, lname FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()
    cur.execute("UPDATE users SET archived_at = NOW() WHERE id = %s", (user_id,))
    conn.commit()
    cur.close()
    conn.close()
    name = f"{user['fname']} {user['lname']}" if user else "User"
    audit_log(session.get("user_id"), "archive_user", "user", user_id, f"Archived {name}")
    flash(f"{name} has been archived.", "warning")
    return redirect("/admin#users")


@admin_bp.route("/admin/user/unarchive/<int:user_id>", methods=["POST"])
def admin_unarchive_user(user_id):
    conn = get_db()
    cur  = dict_cur(conn)
    cur.execute("SELECT fname, lname FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()
    cur.execute("UPDATE users SET archived_at = NULL WHERE id = %s", (user_id,))
    conn.commit()
    cur.close()
    conn.close()
    name = f"{user['fname']} {user['lname']}" if user else "User"
    audit_log(session.get("user_id"), "unarchive_user", "user", user_id, f"Restored {name}")
    flash(f"{name} has been restored.", "success")
    return redirect("/admin#users")


@admin_bp.route("/admin/user/delete/<int:user_id>", methods=["POST"])
def admin_delete_user(user_id):
    conn = get_db()
    cur  = dict_cur(conn)
    cur.execute("SELECT fname, lname, archived_at FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()
    if not user or user["archived_at"] is None:
        cur.close(); conn.close()
        flash("Only archived users can be deleted.", "error")
        return redirect("/admin#users")
    cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
    conn.commit()
    cur.close(); conn.close()
    name = f"{user['fname']} {user['lname']}"
    audit_log(session.get("user_id"), "delete_user", "user", user_id, f"Permanently deleted {name}")
    flash(f"{name} has been permanently deleted.", "error")
    return redirect("/admin#users")


@admin_bp.route("/admin/registration/<reg_id>/status/<status>", methods=["POST"])
def admin_registration_status(reg_id, status):
    if status not in ("accepted", "rejected", "pending"):
        return redirect("/admin#registrations")

    is_qualified = request.form.get("qualified", "1") == "1"
    set_registration_status(reg_id, status, is_qualified if status == "accepted" else None)

    if status == "accepted":
        # Fetch the registration to get email and name
        conn = get_db()
        cur  = dict_cur(conn)
        cur.execute("SELECT email, full_name FROM registrations WHERE id = %s", (reg_id,))
        reg = cur.fetchone()
        cur.close(); conn.close()

        if reg and reg.get("email"):
            try:
                token    = create_account_token(reg_id, reg["email"], is_qualified)
                base_url = request.host_url
                send_acceptance_email(reg["email"], reg["full_name"], token, is_qualified, base_url)
                flash(f"Registration #{reg_id} accepted. Account setup email sent to {reg['email']}.", "success")
            except Exception as e:
                flash(f"Registration #{reg_id} accepted, but email failed: {e}", "warning")
        else:
            flash(f"Registration #{reg_id} accepted. (No email on file — send link manually.)", "warning")
        audit_log(session.get("user_id"), "registration_accepted", "registration", reg_id,
                  f"Accepted registration #{reg_id} (qualified={is_qualified})")
    else:
        audit_log(session.get("user_id"), f"registration_{status}", "registration", reg_id,
                  f"Marked registration #{reg_id} as {status}")
        flash(f"Registration #{reg_id} marked as {status}.", "warning")

    return redirect("/admin#registrations")


@admin_bp.route("/admin/schedule/delete/<int:schedule_id>", methods=["POST"])
def admin_delete_schedule(schedule_id):
    conn = get_db()
    cur  = dict_cur(conn)
    cur.execute("SELECT name FROM batches WHERE id = %s", (schedule_id,))
    row = cur.fetchone()
    cur.execute("DELETE FROM batches WHERE id = %s", (schedule_id,))
    conn.commit()
    cur.close()
    conn.close()
    name = row["name"] if row else "Batch"
    flash(f'"{name}" has been deleted.', "error")
    return redirect("/admin#schedules")


@admin_bp.route("/admin/schedule/add", methods=["POST"])
def admin_add_schedule():
    name      = request.form.get("name",       "").strip()
    start_str = request.form.get("start_date", "").strip()
    btype = request.form.get("btype", "").strip()
    if name and btype in ("training", "renewal") and start_str:
        start_date = date.fromisoformat(start_str)
        offsets    = TYPE_OFFSETS["training"] if btype == "training" else TYPE_OFFSETS["renewal"]
        label      = build_batch_label(name, offsets, start_date)
        conn = get_db()
        cur  = conn.cursor()
        cur.execute("INSERT INTO batches (name, type, start_date) VALUES (%s, %s, %s)", (name, btype, start_date))
        conn.commit()
        cur.close()
        conn.close()
        flash(f'"{label}" has been added.', "success")
    return redirect("/admin#schedules")


# ── Analytics JSON ────────────────────────────────────────────────────────────

@admin_bp.route("/admin/analytics")
def admin_analytics():
    if session.get("user_role") != "admin":
        return jsonify(error="Unauthorized"), 403
    data = get_analytics_data()
    # Convert Decimal to float for JSON serialisation
    for row in data.get("revenue_monthly", []):
        if row.get("revenue") is not None:
            row["revenue"] = float(row["revenue"])
    return jsonify(data)


# ── Excel exports ─────────────────────────────────────────────────────────────

def _xl_response(wb, filename):
    from openpyxl.styles import Font, PatternFill, Alignment
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name=filename)


def _xl_header(ws, headers):
    from openpyxl.styles import Font, PatternFill, Alignment
    fill = PatternFill("solid", fgColor="0D3B27")
    font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill
        c.font = font
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


@admin_bp.route("/admin/export/registrations")
def export_registrations():
    if session.get("user_role") != "admin":
        return redirect("/auth")
    from openpyxl import Workbook
    regs = get_registrations()
    wb = Workbook(); ws = wb.active; ws.title = "Registrations"
    headers = ["Ref #", "Form Type", "Full Name", "Email", "Contact",
               "Company", "Training Type", "Status", "Submitted At"]
    _xl_header(ws, headers)
    for r in regs:
        ws.append([r.get("id"), r.get("form_type"), r.get("full_name"),
                   r.get("email"), r.get("contact_number"), r.get("company_name"),
                   r.get("training_type"), r.get("status"), str(r.get("submitted_at", ""))])
    audit_log(session.get("user_id"), "export_report", "report", None, "Downloaded registrations.xlsx")
    return _xl_response(wb, "registrations.xlsx")


@admin_bp.route("/admin/export/users")
def export_users():
    if session.get("user_role") != "admin":
        return redirect("/auth")
    from openpyxl import Workbook
    users = get_users()
    wb = Workbook(); ws = wb.active; ws.title = "Users"
    headers = ["ID", "Role", "First Name", "Last Name", "Email",
               "Contact", "Company", "Payment Status", "Created At"]
    _xl_header(ws, headers)
    for u in users:
        ws.append([u.get("id"), u.get("role"), u.get("fname"), u.get("lname"),
                   u.get("email"), u.get("contact_number"), u.get("company_name"),
                   u.get("payment_status"), str(u.get("created_at", ""))])
    audit_log(session.get("user_id"), "export_report", "report", None, "Downloaded users.xlsx")
    return _xl_response(wb, "users.xlsx")


@admin_bp.route("/admin/export/payments")
def export_payments():
    if session.get("user_role") != "admin":
        return redirect("/auth")
    from openpyxl import Workbook
    conn = get_db(); cur = dict_cur(conn)
    cur.execute("""
        SELECT p.id, u.fname, u.lname, u.email, p.amount_php, p.method,
               p.status, p.receipt_id, p.tx_hash, p.created_at, p.paid_at
        FROM payments p JOIN users u ON p.user_id = u.id ORDER BY p.created_at DESC
    """)
    payments = cur.fetchall(); cur.close(); conn.close()
    wb = Workbook(); ws = wb.active; ws.title = "Payments"
    headers = ["ID", "First Name", "Last Name", "Email", "Amount (PHP)", "Method",
               "Status", "Receipt ID", "TX Hash", "Created At", "Paid At"]
    _xl_header(ws, headers)
    for p in payments:
        ws.append([p["id"], p["fname"], p["lname"], p["email"],
                   float(p["amount_php"]) if p["amount_php"] else 0,
                   p["method"], p["status"], p.get("receipt_id"),
                   p.get("tx_hash"), str(p["created_at"]), str(p.get("paid_at") or "")])
    audit_log(session.get("user_id"), "export_report", "report", None, "Downloaded payments.xlsx")
    return _xl_response(wb, "payments.xlsx")


@admin_bp.route("/admin/export/lms-progress")
def export_lms_progress():
    if session.get("user_role") != "admin":
        return redirect("/auth")
    from openpyxl import Workbook
    conn = get_db(); cur = dict_cur(conn)
    cur.execute("""
        SELECT u.fname, u.lname, u.email, m.title AS module,
               p.quiz_score, p.quiz_passed, p.exam_grade,
               p.completed, p.completed_at
        FROM lms_progress p
        JOIN users u ON p.user_id = u.id
        JOIN lms_modules m ON p.module_id = m.id
        ORDER BY u.lname, m.title
    """)
    rows = cur.fetchall(); cur.close(); conn.close()
    wb = Workbook(); ws = wb.active; ws.title = "LMS Progress"
    headers = ["First Name", "Last Name", "Email", "Module",
               "Quiz Score", "Quiz Passed", "Exam Grade", "Completed", "Completed At"]
    _xl_header(ws, headers)
    for r in rows:
        ws.append([r["fname"], r["lname"], r["email"], r["module"],
                   float(r["quiz_score"]) if r["quiz_score"] else None,
                   "Yes" if r["quiz_passed"] else "No",
                   r["exam_grade"],
                   "Yes" if r["completed"] else "No",
                   str(r["completed_at"] or "")])
    audit_log(session.get("user_id"), "export_report", "report", None, "Downloaded lms-progress.xlsx")
    return _xl_response(wb, "lms-progress.xlsx")


@admin_bp.route("/admin/export/solar-requests")
def export_solar_requests():
    if session.get("user_role") != "admin":
        return redirect("/auth")
    from openpyxl import Workbook
    reqs = get_solar_requests()
    wb = Workbook(); ws = wb.active; ws.title = "Solar Requests"
    headers = ["ID", "Name", "Email", "Contact", "Establishment",
               "Monthly Bill (PHP)", "System Size (kW)", "Status", "Created At"]
    _xl_header(ws, headers)
    for r in reqs:
        ws.append([r.get("id"), r.get("name"), r.get("email"), r.get("contact"),
                   r.get("establishment_type"),
                   float(r.get("monthly_bill_php") or 0),
                   float(r.get("system_size_kw") or 0),
                   r.get("status"), str(r.get("created_at", ""))])
    audit_log(session.get("user_id"), "export_report", "report", None, "Downloaded solar-requests.xlsx")
    return _xl_response(wb, "solar-requests.xlsx")


# ── Feedback JSON ─────────────────────────────────────────────────────────────

@admin_bp.route("/admin/feedback-json")
def admin_feedback_json():
    if session.get("user_role") != "admin":
        return jsonify(error="Unauthorized"), 403
    feedback = get_feedback()
    summary  = get_feedback_summary()
    # Serialise datetime objects
    for row in feedback:
        if row.get("created_at"):
            row["created_at"] = str(row["created_at"])
    for s in summary:
        s["module_title"] = s["title"]
        s["avg_rating"] = float(s["avg_rating"] or 0)
    return jsonify(feedback=feedback, summary=summary)


# ── Audit log JSON ────────────────────────────────────────────────────────────

@admin_bp.route("/admin/audit-log-json")
def admin_audit_log_json():
    if session.get("user_role") != "admin":
        return jsonify(error="Unauthorized"), 403
    logs = get_audit_logs(limit=500)
    for row in logs:
        if row.get("created_at"):
            row["created_at"] = str(row["created_at"])
    return jsonify(logs=logs)

